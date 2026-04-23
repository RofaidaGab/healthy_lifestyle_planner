import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook, Workbook
import random
import json


class HealthyLifestyleApp:
    def __init__(self, master: tk.Tk):
        self.master = master
        self.master.title("Healthy Lifestyle Planner")
        self.master.geometry("440x600+400+50")

        # Data
        self.file_path = ""
        self.categories = {}
        self.chosen_sports = []
        self.chosen_food = []
        self.chosen_activities = []

        self.extensions = {
            "TXT": ".txt",
            "JSON": ".json",
            "Excel": ".xlsx"
        }

        self._build_ui()

    # ---------------- UI ---------------- #
    def _build_ui(self):
        ttk.Button(self.master, text="Import Excel File",
                   command=self.import_excel).grid(row=0, column=1, pady=5)

        # Category selector
        self.category_var = tk.StringVar()
        self.category_box = ttk.Combobox(
            self.master,
            textvariable=self.category_var,
            state="readonly"
        )
        self.category_box.grid(row=1, column=0, padx=15, pady=10)
        self.category_box.bind("<<ComboboxSelected>>", self.update_listbox)

        # Listboxes
        self.available_listbox = tk.Listbox(self.master)
        self.available_listbox.grid(row=2, column=0, padx=6, pady=10)

        self.selected_listbox = tk.Listbox(self.master)
        self.selected_listbox.grid(row=2, column=2, padx=6, pady=10)

        # Buttons
        ttk.Button(self.master, text="Add", command=self.add_item)\
            .grid(row=3, column=0, pady=2)

        ttk.Button(self.master, text="Delete", command=self.delete_item)\
            .grid(row=4, column=0, pady=2)

        ttk.Button(self.master, text="Randomize", command=self.randomize)\
            .grid(row=5, column=0, pady=2)

        # Output text
        self.textbox = tk.Text(self.master, width=50, height=8)
        self.textbox.grid(row=6, column=0, columnspan=3, padx=10, pady=10)
        self._reset_textbox()

        # Week input
        tk.Label(self.master, text="Week Number:").grid(row=7, column=0)
        self.week_entry = ttk.Entry(self.master)
        self.week_entry.grid(row=7, column=2)

        # Export
        self.export_var = tk.StringVar(value="TXT")
        ttk.Combobox(self.master,
                     values=list(self.extensions.keys()),
                     textvariable=self.export_var,
                     state="readonly").grid(row=8, column=0)

        ttk.Button(self.master, text="Export",
                   command=self.export).grid(row=8, column=2)

    # ---------------- Core Logic ---------------- #
    def add_item(self):
        idx = self.available_listbox.curselection()
        if not idx:
            messagebox.showwarning("No Selection", "Please select an item!")
            return

        item = self.available_listbox.get(idx)
        category = self.category_var.get()

        target = self._get_category_list(category)

        if len(target) >= 7:
            messagebox.showwarning("Limit Reached", "Max 7 items allowed.")
            return

        if item not in target:
            target.append(item)
            self.selected_listbox.insert(tk.END, item)
            self.update_textbox()

    def delete_item(self):
        idx = self.selected_listbox.curselection()
        if not idx:
            messagebox.showwarning("No Selection", "Please select an item!")
            return

        item = self.selected_listbox.get(idx)
        category = self.category_var.get()

        target = self._get_category_list(category)

        if item in target:
            target.remove(item)

        self.selected_listbox.delete(idx)
        self.update_textbox()

    def randomize(self):
        self._clear_selected()

        self.chosen_sports = self._random_pick("Sports")
        self.chosen_food = self._random_pick("Food Plans")
        self.chosen_activities = self._random_pick("Activities")

        for item in self._current_selected():
            self.selected_listbox.insert(tk.END, item)

        self.update_textbox()

    # ---------------- Textbox ---------------- #
    def update_textbox(self):
        self.textbox.delete("1.0", tk.END)

        self.textbox.insert(tk.END, "Sports:\n" + "\n".join(self.chosen_sports))
        self.textbox.insert(tk.END, "\n\nFood Plans:\n" + "\n".join(self.chosen_food))
        self.textbox.insert(tk.END, "\n\nActivities:\n" + "\n".join(self.chosen_activities))

    def _reset_textbox(self):
        self.textbox.delete("1.0", tk.END)
        self.textbox.insert(tk.END, "Sports:\n\nFood Plans:\n\nActivities:\n")

    # ---------------- Excel / Import ---------------- #
    def import_excel(self):
        file = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not file:
            return

        try:
            self.categories = self._load_excel(file)
            self._update_combobox()
            messagebox.showinfo("Success", "File imported successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _load_excel(self, file):
        wb = load_workbook(file)
        sheet = wb.active

        letters = ["A", "B", "C"]
        self.categories.clear()

        for col, letter in enumerate(letters):
            key = sheet[f"{letter}1"].value
            if not key:
                continue

            self.categories[key] = []

            for row in range(2, 28):
                val = sheet[f"{letter}{row}"].value
                if val:
                    self.categories[key].append(val)

        return self.categories

    def _update_combobox(self):
        keys = list(self.categories.keys())
        self.category_box["values"] = keys

        if keys:
            self.category_var.set(keys[0])

        self.update_listbox()

    # ---------------- Listbox ---------------- #
    def update_listbox(self, event=None):
        self.available_listbox.delete(0, tk.END)

        category = self.category_var.get()
        for item in self.categories.get(category, []):
            self.available_listbox.insert(tk.END, item)

        self.selected_listbox.delete(0, tk.END)

    # ---------------- Export ---------------- #
    def export(self):
        week = self.week_entry.get()
        ext = self.export_var.get()

        if not week.isdigit():
            messagebox.showerror("Error", "Enter valid week number.")
            return

        name = f"HealthyLifeStyle_Week{week}"

        try:
            if ext == "TXT":
                with open(f"{name}.txt", "w") as f:
                    f.write("Sports:\n" + "\n".join(self.chosen_sports))
                    f.write("\n\nFood Plans:\n" + "\n".join(self.chosen_food))
                    f.write("\n\nActivities:\n" + "\n".join(self.chosen_activities))

            elif ext == "JSON":
                with open(f"{name}.json", "w") as f:
                    json.dump({
                        "Sports": self.chosen_sports,
                        "Food Plans": self.chosen_food,
                        "Activities": self.chosen_activities
                    }, f, indent=4)

            elif ext == "Excel":
                wb = Workbook()
                ws = wb.active
                ws.title = "Plan"

                ws.append(["Day", "Sport", "Food", "Activity"])

                for i in range(7):
                    ws.append([
                        f"Day {i+1}",
                        self.chosen_sports[i] if i < len(self.chosen_sports) else "",
                        self.chosen_food[i] if i < len(self.chosen_food) else "",
                        self.chosen_activities[i] if i < len(self.chosen_activities) else "",
                    ])

                wb.save(f"{name}.xlsx")

            messagebox.showinfo("Success", "Export completed!")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ---------------- Helpers ---------------- #
    def _get_category_list(self, category):
        if category == "Sports":
            return self.chosen_sports
        if category == "Food Plans":
            return self.chosen_food
        if category == "Activities":
            return self.chosen_activities
        return []

    def _random_pick(self, category):
        items = self.categories.get(category, [])
        return random.sample(items, min(7, len(items)))

    def _current_selected(self):
        return self.chosen_sports + self.chosen_food + self.chosen_activities

    def _clear_selected(self):
        self.selected_listbox.delete(0, tk.END)
        self.chosen_sports.clear()
        self.chosen_food.clear()
        self.chosen_activities.clear()


def main():
    root = tk.Tk()
    app = HealthyLifestyleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()